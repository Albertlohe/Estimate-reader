import re
import io
import logging
from contextlib import asynccontextmanager
from pathlib import Path

from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

logger = logging.getLogger("npwd_sor")

# ── Bundled SOR JSON (generated once locally by build_sor_json.py) ──────────
BUNDLED_SOR_JSON = Path(__file__).parent / "sor_data.json"


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Auto-load pre-parsed SOR data from JSON on startup (fast, low RAM)."""
    global sor_data, sor_loaded
    if BUNDLED_SOR_JSON.exists():
        try:
            import json as _json
            with open(BUNDLED_SOR_JSON, encoding="utf-8") as f:
                merged = _json.load(f)
            if merged:
                sor_data   = merged
                sor_loaded = True
                mf_count   = sum(1 for v in merged.values() if v["has_mf"])
                logger.info(
                    f"✅ SOR loaded from JSON — "
                    f"{len(merged)} items ({mf_count} MF-flagged) across 12 districts."
                )
            else:
                logger.warning("⚠️  sor_data.json is empty.")
        except Exception as exc:
            logger.error(f"❌ Failed to load sor_data.json: {exc}")
    else:
        logger.warning(
            "⚠️  sor_data.json not found. Run build_sor_json.py locally first, "
            "then commit the JSON file to GitHub."
        )
    yield  # app runs here


app = FastAPI(title="NPWD SOR Rate Lookup Tool", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

DISTRICTS = ["Kohima","Dimapur","Peren","Wokha","Phek",
             "Zunheboto","Mokokchung","Tuensang","Mon",
             "Longleng","Kiphire","Noklak"]

BLOCK_MF = {
    "Kohima":     {"Kohima HQ":1.00,"ADC Tseminyu":1.04,"SDO(C) Sechu":0.99,"SDO(C) Chiephobozu":1.02,"SDO(C) Jakhama":1.01},
    "Peren":      {"Peren HQ":1.00,"SDO(C) Jalukie":0.99,"SDO(C) Tening":1.04},
    "Phek":       {"Phek HQ":1.00,"ADC Pfutsero":0.93,"ADC Chozouba":0.93,"ADC Meluri":1.01,"SDO(C) Chizami":0.94},
    "Kiphire":    {"Kiphire HQ":1.00,"SDO(C) Seyochung":1.01,"SDO(C) Pungro":1.02},
    "Longleng":   {"Longleng HQ":1.00,"SDO(C) Tamlu":0.97},
    "Dimapur":    {"Dimapur HQ":1.00,"ADC Niuland":1.02,"ADC Medziphema":1.02,"SDO(C) Dhansiripar":1.02,"SDO(C) Kuhuboto":1.01},
    "Wokha":      {"Wokha HQ":1.00,"ADC Bhandari":0.92,"ADC DHEP":0.93,"SDO(C) Ralan":0.90,"SDO(C) Sanis":0.93},
    "Mokokchung": {"Mokokchung HQ":1.00,"ADC Mangkolemba":0.97,"ADC Tuli":0.96,"SDO(C) Changtongya":0.98},
    "Tuensang":   {"Tuensang HQ":1.00,"ADC Shamatore":1.06,"ADC Longkhim":0.96},
    "Zunheboto":  {"Zunheboto HQ":1.00,"ADC Aghunato":1.06,"ADC Phughoboto":0.90,"SDO(C) Suruhoto":0.98,"SDO(C) Akuluto":0.95,"SDO(C) Atoizu":0.98,"SDO(C) Satakha":0.98},
    "Mon":        {"Mon HQ":1.00,"ADC Aboi":1.02,"ADC Naginimora":0.95,"ADC Tizit":0.97,"ADC Tobu":1.25,"SDO(C) Wakching":1.04,"SDO(C) Phomching":1.02,"SDO(C) Chen":1.12,"SDO(C) Angjangyang":1.12,"SDO(C) Manyakshu":1.25},
    "Noklak":     {"Noklak HQ":1.00,"SDO(C) Thonoknyu":1.06},
}

# In-memory SOR store
# { norm_code -> { description, unit, rates:{district->float}, has_mf:bool } }
sor_data: dict = {}
sor_loaded: bool = False


def normalize_code(code: str) -> str:
    code = code.strip()
    code = re.sub(r'\s*\.\s*', '.', code)
    code = re.sub(r'^A\s+', 'A ', code)
    code = code.rstrip('.')
    return code.upper()


def parse_sor_text(pdf_bytes: bytes) -> dict:
    result = {}
    UNITS = ['Cum','Sqm','Kg','Metre','Each','Litre','each','metre',
             'Sqft','Quintal','Tonne','Nos','Running Metre','RM','Hour','Set']
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                line = line.strip()
                m = re.match(r'^(A\s*\d+(?:\.\d+[A-Za-z]?)*)\s*(.*)', line)
                if not m:
                    continue
                raw_code = m.group(1)
                rest     = m.group(2).strip()
                norm     = normalize_code(raw_code)
                has_mf   = bool(re.match(r'^MF\b', rest))
                rest     = re.sub(r'^MF\s*', '', rest).strip()
                all_num_matches = list(re.finditer(r'\b\d{1,7}(?:\.\d+)?\b', rest))
                if len(all_num_matches) < 12:
                    continue
                rate_nums = [m.group() for m in all_num_matches[-12:]]
                # Everything before the first of the last-12 numbers is desc+unit
                rate_block_start = all_num_matches[-12].start()
                desc_part = rest[:rate_block_start].strip()
                unit = ""
                for u in UNITS:
                    if desc_part.endswith(u):
                        unit = u; desc_part = desc_part[:-len(u)].strip(); break
                    if desc_part == u:
                        unit = u; desc_part = ""; break
                desc_part = re.sub(r'\s+', ' ', desc_part)
                try:
                    rates = {DISTRICTS[k]: float(rate_nums[k]) for k in range(12)}
                except (ValueError, IndexError):
                    continue
                if norm not in result:
                    result[norm] = {"description": desc_part, "unit": unit, "rates": rates, "has_mf": has_mf}
                else:
                    if len(rates) > len(result[norm]["rates"]): result[norm]["rates"] = rates
                    if has_mf: result[norm]["has_mf"] = True
                    if not result[norm]["unit"] and unit: result[norm]["unit"] = unit
    return result


def parse_sor_table(pdf_bytes: bytes) -> dict:
    result = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if not tables: continue
            for table in tables:
                if not table: continue
                header_row = None; header_idx = None
                for i, row in enumerate(table):
                    if row and any(str(c or '').strip() in ("Kohima","Code No.","Code No") for c in row):
                        header_row = row; header_idx = i; break
                if header_row is None: continue
                cells = [str(c or '').strip() for c in header_row]
                try:
                    code_idx = next(i for i,c in enumerate(cells) if "Code" in c)
                except StopIteration:
                    continue
                desc_idx = next((i for i,c in enumerate(cells) if "Desc" in c), None)
                unit_idx = next((i for i,c in enumerate(cells) if c=="Unit"), None)
                mf_idx   = next((i for i,c in enumerate(cells) if c=="MF"), None)
                dist_indices = {}
                for d in DISTRICTS:
                    for i,c in enumerate(cells):
                        if d in c: dist_indices[d] = i; break
                if not dist_indices: continue
                current_code = None; current_desc = ""; current_unit = ""; current_has_mf = False
                for row in table[header_idx+1:]:
                    if not row: continue
                    rc = [str(c or '').strip() for c in row]
                    code_cell = rc[code_idx] if code_idx < len(rc) else ""
                    desc_cell = rc[desc_idx] if desc_idx is not None and desc_idx < len(rc) else ""
                    unit_cell = rc[unit_idx] if unit_idx is not None and unit_idx < len(rc) else ""
                    mf_cell   = rc[mf_idx]   if mf_idx is not None and mf_idx < len(rc) else ""
                    cm = re.match(r'^(A\s*\d[\d\s.]*[A-Za-z]?[\d.]*)', code_cell)
                    if cm:
                        current_code   = normalize_code(cm.group(1))
                        current_desc = " ".join(desc_cell.split())
                        current_unit   = unit_cell
                        current_has_mf = (mf_cell.strip().upper() == "MF")
                    else:
                        if current_code and desc_cell:
                            current_desc = " ".join((current_desc + " " + desc_cell).split())

                            # keep updating stored description
                            if current_code in result:
                                result[current_code]["description"] = current_desc

                        if current_code and unit_cell:
                            current_unit = unit_cell

                            if current_code in result:
                                result[current_code]["unit"] = current_unit
                    if current_code:
                        if current_code not in result:
                            result[current_code] = {"description": current_desc, "unit": current_unit,
                                                    "rates": {}, "has_mf": current_has_mf}
                        for dist, di in dist_indices.items():
                            if di < len(rc):
                                try:
                                    v = float(rc[di].replace(",",""))
                                    if v > 0: result[current_code]["rates"][dist] = v
                                except ValueError: pass
    return result


@app.get("/", response_class=HTMLResponse)
async def root():
    return (Path(__file__).parent / "index.html").read_text(encoding="utf-8")


@app.post("/upload-sor")
async def upload_sor(file: UploadFile = File(...)):
    global sor_data, sor_loaded
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files accepted.")
    content = await file.read()
    # Use BOTH parsers
    table_data = parse_sor_table(content)
    text_data = parse_sor_text(content)

    merged = dict(table_data)

    for code, item in text_data.items():
        # Add items missed by table parser
        if code not in merged:
            merged[code] = item
        else:
            # Keep table rates (IMPORTANT)
            if not merged[code]["description"].strip() and item["description"].strip():
                merged[code]["description"] = item["description"]

            if not merged[code]["unit"] and item["unit"]:
                merged[code]["unit"] = item["unit"]

            if item["has_mf"]:
                merged[code]["has_mf"] = True

    if not merged:
        raise HTTPException(422, "Could not extract items.")

    sor_data = merged
    sor_loaded = True

    mf_count = sum(1 for v in merged.values() if v["has_mf"])
    return {"success": True, "items_loaded": len(merged), "mf_items": mf_count,
            "message": f"SOR loaded — {len(merged)} items ({mf_count} MF-flagged) across 12 districts."}


@app.get("/sor-status")
async def sor_status():
    return {"loaded": sor_loaded, "items": len(sor_data), "districts": DISTRICTS}


@app.get("/block-mf")
async def get_block_mf():
    return BLOCK_MF


# ── CORE LOOKUP ──────────────────────────────────────────────────────────────
# INPUT:  code, district, block_mf (the Annexure-I factor, e.g. 1.02)
# LOGIC:  if item has_mf → final = base_rate × block_mf
#         if item NOT has_mf → final = base_rate × 1.00  (block_mf ignored)
# OUTPUT: base_rate, has_mf, mf_applied (1.02 or 1.00), final_rate
# ─────────────────────────────────────────────────────────────────────────────
class LookupRequest(BaseModel):
    codes: list[dict]   # [{code, district, block_mf}]


@app.post("/lookup")
async def lookup_rates(req: LookupRequest):
    if not sor_loaded:
        raise HTTPException(400, "SOR not loaded.")
    results = []
    for item in req.codes:
        raw_code  = item.get("code", "").strip()
        district  = item.get("district", "Kohima").strip()
        block_mf  = float(item.get("block_mf", 1.0))   # e.g. 1.02 from Annexure-I

        norm  = normalize_code(raw_code)
        match = sor_data.get(norm)
        if not match:
            alt = re.sub(r'\s+', '', norm)
            for sk in sor_data:
                if re.sub(r'\s+', '', sk) == alt:
                    match = sor_data[sk]; norm = sk; break

        if match:
            quantity = float(item.get("quantity", 1) or 1)

            base_rate = float(
                match["rates"].get(district)
                or match["rates"].get("Kohima", 0.0)
            )

            has_mf = bool(match["has_mf"])
            selected_mf = round(float(block_mf), 2)

            # MF items use selected Block MF; non-MF items always use 1.00.
            mf_applied = selected_mf if has_mf else 1.00
            final_rate = round(base_rate * mf_applied, 2) if has_mf else round(base_rate, 2)
            amount = round(final_rate * quantity, 2)

            results.append({
                "code": raw_code,
                "description": match["description"],
                "unit": match["unit"],
                "quantity": round(quantity, 2),
                "district": district,
                "base_rate": round(base_rate, 2),
                "has_mf": has_mf,
                "mf_applied": round(mf_applied, 2),
                "final_rate": round(final_rate, 2),
                "amount": round(amount, 2),
                "block_mf": round(selected_mf, 2),
                "found": True
            })

        else:
            results.append({
                "code": raw_code, "description": "NOT FOUND", "unit": "-",
                "quantity": 0, "district": district, "base_rate": 0,
                "has_mf": False, "mf_applied": 1.0, "final_rate": 0,
                "amount": 0, "block_mf": 1.0, "found": False
            })
    return {"results": results,
            "total": len(results),
            "found": sum(1 for r in results if r["found"]),
            "not_found": [r["code"] for r in results if not r["found"]]}


@app.post("/upload-excel-codes")
async def upload_excel_codes(file: UploadFile = File(...),
                             district: str  = Form("Kohima"),
                             block_mf: str  = Form("1.0")):
    if not sor_loaded:
        raise HTTPException(400, "SOR not loaded.")
    bmf = float(block_mf) if block_mf else 1.0
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    codes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        code = str(row[0]).strip()
        quantity = 1.0
        row_district = district
        row_bmf = bmf

        if len(row) > 1 and row[1] is not None:
            try:
                quantity = float(row[1])
            except:
                pass

        if len(row) > 2 and row[2] is not None:
            v = str(row[2]).strip()
            if v in DISTRICTS:
                row_district = v
            else:
                try:
                    row_bmf = float(v)
                except:
                    pass

        if len(row) > 3 and row[3] is not None:
            v = str(row[3]).strip()
            if v in DISTRICTS:
                row_district = v
            else:
                try:
                    row_bmf = float(v)
                except:
                    pass

        codes.append({
            "code": code,
            "quantity": quantity,
            "district": row_district,
            "block_mf": row_bmf
        })

    if not codes:
        raise HTTPException(422, "No codes found. Put codes in Column A from row 2.")
    return await lookup_rates(LookupRequest(codes=codes))



@app.post("/export-excel")
async def export_excel(req: LookupRequest):
    if not sor_loaded:
        raise HTTPException(400, "SOR not loaded.")

    data = await lookup_rates(req)
    results = data["results"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cost Abstract"

    # ── Page layout / styles ────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    sub_fill = PatternFill("solid", fgColor="D9E2F3")
    grey_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    title_font = Font(name="Calibri", bold=True, color="000000", size=12)
    hdr_font = Font(name="Calibri", bold=True, color="000000", size=10)
    body_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", bold=True, size=10)
    red_font = Font(name="Calibri", size=10, color="B71C1C")

    # Column widths close to the sample workbook
    widths = {
        "A": 13.0,
        "B": 13.33,
        "C": 68.55,
        "D": 13.0,
        "E": 13.0,
        "F": 16.66,
        "G": 13.22,
        "H": 20.0,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ── Header block ───────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = "2. Cost Abstract"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = sub_fill
    ws["A1"].border = border
    ws.row_dimensions[1].height = 22

    ws["A2"] = "Name of the Work: Generated Estimate from SOR 2021 Codes."
    ws["A2"].font = body_font
    ws["A2"].alignment = left

    ws["A3"] = "Item of the Work: SOR Rate Abstract"
    ws["A3"].font = body_font
    ws["A3"].alignment = left

    ws["A4"] = "Note:"
    ws["A4"].font = bold_font
    ws["A4"].alignment = left
    ws["C4"] = "Item No. refers to the serial item number of this estimate."
    ws["C4"].font = body_font
    ws["C4"].alignment = left

    ws["C5"] = "Schedule Number refers to the corresponding item number in the Nagaland PWD Schedule of Rates, 2021"
    ws["C5"].font = body_font
    ws["C5"].alignment = left

    headers = [
        "Item No",
        "Schedule Number",
        "Description of Item",
        "Quantity",
        "Unit",
        "Rate",
        "Multiplication factor (MF)",
        "Amount\nin Rupees",
    ]
    header_row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = sub_fill
        cell.border = border
        cell.alignment = center
    ws.row_dimensions[header_row].height = 38

    # Spacer rows
    ws.row_dimensions[7].height = 6
    ws.row_dimensions[8].height = 6

    # Optional section header to mimic the sample layout
    current_row = 9
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    c = ws.cell(row=current_row, column=1, value="SOR ITEMS")
    c.font = bold_font
    c.alignment = center
    c.fill = grey_fill
    c.border = border
    ws.row_dimensions[current_row].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────
    data_start = current_row + 1
    for i, item in enumerate(results, start=1):
        row = data_start + i - 1

        qty = float(item.get("quantity", 1) or 1)
        rate = float(item.get("base_rate", 0) or 0)
        mf = float(item.get("mf_applied", 1) or 1)
        amount = float(item.get("amount", round(rate * qty * mf, 2)) or 0)

        row_values = [
            f"1.{i}",
            item.get("code", ""),
            item.get("description", ""),
            qty,
            item.get("unit", ""),
            rate,
            mf,
            amount,
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = red_font if not item.get("found", True) else body_font
            cell.border = border
            cell.alignment = left if col_idx in (2, 3) else center

        ws.cell(row=row, column=4).number_format = "0.00"
        ws.cell(row=row, column=6).number_format = '₹ #,##0.00'
        ws.cell(row=row, column=7).number_format = "0.00"
        ws.cell(row=row, column=8).number_format = '₹ #,##0.00'
        ws.row_dimensions[row].height = 42

    total_row = data_start + len(results)
    ws.merge_cells(start_row=total_row, start_column=4, end_row=total_row, end_column=7)
    tcell = ws.cell(row=total_row, column=4, value="Total of Estimate:")
    tcell.font = bold_font
    tcell.alignment = right
    tcell.fill = grey_fill
    tcell.border = border

    hcell = ws.cell(row=total_row, column=8)
    if len(results) > 0:
        hcell.value = f"=SUM(H{data_start}:H{total_row-1})"
    else:
        hcell.value = 0
    hcell.font = bold_font
    hcell.alignment = center
    hcell.fill = grey_fill
    hcell.border = border
    hcell.number_format = '₹ #,##0.00'
    ws.row_dimensions[total_row].height = 22

    ws.freeze_panes = "A7"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Cost_Abstract.xlsx"},
    )


@app.get("/search")
async def search_codes(q: str):
    if not sor_loaded: return {"results": []}
    q_lower = q.lower().strip()
    matches = []
    for code, data in sor_data.items():
        if q_lower in code.lower() or q_lower in data["description"].lower():
            matches.append({"code": code, "description": data["description"][:100],
                            "unit": data["unit"], "has_mf": data["has_mf"]})
        if len(matches) >= 20: break
    return {"results": matches}
